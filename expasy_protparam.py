from selenium import webdriver
from selenium.webdriver.common.by import By
import time
import re
from openpyxl import Workbook, load_workbook

# === Step 1: Load accession numbers and lengths from Excel ===

excel_file = "C:\\Users\\Santhoshi\\PycharmProjects\\Project1\\Python_Practise\\uniprotkb_human_uncharacterized_protein.xlsx"
workbook = load_workbook(excel_file)
sheet = workbook.active

accession_list = []  # List of accession numbers
uniprot_lengths = {}  # Dict of accession: length

# Assuming 'entry' is in column A (index 0) and 'length' is in column G (index 6)
for row in sheet.iter_rows(min_row=2, values_only=True):  # Skip header
    if row[0] and row[6]:
        accession = str(row[0]).strip()
        length_str = str(row[6]).strip()
        length = int(''.join(filter(str.isdigit, length_str)))  # Remove " AA" or other text
        accession_list.append(accession)
        uniprot_lengths[accession] = length

# === Step 2: Setup output workbook ===

wb = Workbook()
ws = wb.active
ws.title = "ProtParam Results"
ws.append(["Accession", "Label", "Value"])

# === Step 3: Start Selenium and loop through accessions ===

driver = webdriver.Chrome()
driver.maximize_window()

for accession in accession_list:
    print(f"🔄 Processing {accession}...")
    try:
        driver.get("https://web.expasy.org/protparam/")
        time.sleep(2)

        text_box = driver.find_element(By.XPATH, "/html/body/main/div/form/textarea")
        text_box.clear()
        text_box.send_keys(accession)
        time.sleep(2)

        click_parameters = driver.find_element(By.XPATH, "/html/body/main/div/form/input[3]")
        click_parameters.click()
        time.sleep(2)

        element_1 = driver.find_element(By.CSS_SELECTOR, "body > main > div > pre > strong > a")
        element_1.click()
        time.sleep(2)

        exclude_labels = {
            "Formula:",
            "Extinction coefficients:",
            "Estimated half-life:",
            "Amino acid composition:",
            "Atomic composition:"
        }

        strong_elements = driver.find_elements(By.CSS_SELECTOR, "pre strong")

        for strong in strong_elements:
            label = strong.text.strip()
            if label in exclude_labels:
                continue

            parent = strong.find_element(By.XPATH, './parent::*')
            full_text = parent.text.strip()
            label_index = full_text.find(label)
            after_label = full_text[label_index + len(label):].strip()

            match = re.search(r'-?\d+(\.\d+)?', after_label)
            number = match.group() if match else "N/A"

            ws.append([accession, label, number])

    except Exception as e:
        print(f"❌ Error processing {accession}: {e}")
        ws.append([accession, "ERROR", str(e)])

# === Step 4: Compare UniProt vs Expasy lengths ===

ws2 = wb.create_sheet(title="Signal Comparison")
ws2.append(["Accession", "UniProt Length", "Expasy Length", "Difference", "Signal Info"])

for accession in accession_list:
    uni_len = uniprot_lengths.get(accession)
    exp_len = None

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == accession and row[1] == "Number of amino acids:":
            try:
                exp_len = int(row[2])
            except:
                exp_len = None
            break

    if exp_len is not None:
        diff = uni_len - exp_len
        signal = "No signal" if diff == 0 else "Presence of signal"
        ws2.append([accession, uni_len, exp_len, diff, signal])
    else:
        ws2.append([accession, uni_len, "N/A", "N/A", "Expasy value missing"])

# === Step 6: pI Classification ===

ws3 = wb.create_sheet(title="PI analysis")
ws3.append(["Accession", "Theoretical pI", "Nature"])

for row in ws.iter_rows(min_row=2, values_only=True):
    accession, label, value = row
    if label == "Theoretical pI:":
        try:
            pi = float(value)
            if pi < 7:
                nature = "Acidic"
            elif pi == 7:
                nature = "Neutral"
            else:
                nature = "Basic"
            ws3.append([accession, pi, nature])
        except ValueError:
            ws3.append([accession, value, "Invalid pI"])


# === Step 7: GRAVY Classification ===

ws4 = wb.create_sheet(title="GRAVY analysis")
ws4.append(["Accession", "GRAVY Value", "Nature"])

for row in ws.iter_rows(min_row=2, values_only=True):
    accession, label, value = row
    if label == "Grand average of hydropathicity (GRAVY):":
        try:
            gravy = float(value)
            if gravy < 0:
                nature = "Hydrophilic"
            else:
                nature = "Hydrophobic"
            ws4.append([accession, gravy, nature])
        except ValueError:
            ws4.append([accession, value, "Invalid GRAVY"])

# === Step 8: Instability Index Classification ===

ws5 = wb.create_sheet(title="Instability analysis")
ws5.append(["Accession", "Instability Index", "Stability"])

for row in ws.iter_rows(min_row=2, values_only=True):
    accession, label, value = row
    if label == "Instability index:":
        try:
            index = float(value)
            if index < 40:
                stability = "Stable"
            else:
                stability = "Unstable"
            ws5.append([accession, index, stability])
        except ValueError:
            ws5.append([accession, value, "Invalid Index"])

# === Step 9: Aliphatic Index Thermostability Analysis ===

ws6 = wb.create_sheet(title="Aliphatic index analysis")
ws6.append(["Accession", "Aliphatic Index", "Thermostability"])

for row in ws.iter_rows(min_row=2, values_only=True):
    accession, label, value = row
    if label == "Aliphatic index:":
        try:
            aliph_index = float(value)
            if aliph_index > 100:
                stability = "Very good thermostability"
            elif 80 <= aliph_index <= 100:
                stability = "Good thermostability"
            else:
                stability = "Bad thermostability"

            ws6.append([accession, aliph_index, stability])
        except ValueError:
            ws6.append([accession, value, "Invalid Index"])

# === Step 11: Molecular Weight Sheet ===

ws8 = wb.create_sheet(title="Molecular Weight")
ws8.append(["Accession", "Molecular Weight"])

for row in ws.iter_rows(min_row=2, values_only=True):
    accession, label, value = row
    if label == "Molecular weight:":
        ws8.append([accession, value])

# === Step 12: Charged Residues Sheet ===

ws9 = wb.create_sheet(title="Charged Residues")
ws9.append(["Accession", "Total Positively Charged Residues", "Total Negatively Charged Residues"])

for accession in accession_list:
    pos_res = None
    neg_res = None

    # Search all rows to find these values for the current accession
    for row in ws.iter_rows(min_row=2, values_only=True):
        acc, label, value = row
        if acc == accession:
            if label == "Total number of positively charged residues (Arg + Lys):":
                try:
                    pos_res = int(value)
                except:
                    pos_res = None
            elif label == "Total number of negatively charged residues (Asp + Glu):":
                try:
                    neg_res = int(value)
                except:
                    neg_res = None
        # If both found, no need to keep searching
        if pos_res is not None and neg_res is not None:
            break

    ws9.append([
        accession,
        pos_res if pos_res is not None else "N/A",
        neg_res if neg_res is not None else "N/A"
    ])



# === Step 13: Save and quit ===

wb.save("all_proteins_results.xlsx")
driver.quit()
print("✅ All results saved to all_proteins_results.xlsx")
